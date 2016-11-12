from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Alignment

from lxml.html import fromstring
from analysis  import get_weighted_scores
from utils     import login, pdf_word_count
from string    import uppercase as ALPHABET 

import datetime
import os
import json
import requests


BASE_URL = 'https://peerfeedback.gatech.edu'
GRADING_TEMPLATE_PATH = 'templates/KBAI PF Grading Template.xltx'

# Peer Feedback Site Xpaths
RUBRIC_TEXT_XPATH       = '//div[@class="rubricContainerOpen"]//td[@class="rubricNoSelect"]/div'
ASSIGNMENT_NAME_XPATH   = '//div[contains(@class, "taskCard")]//h4'
ASSIGNMENT_LINK_XPATH   = '//h2/a[contains(.,"Download submitted assignment")]'
ASSIGNMENT_LINKS_XPATH  = '//a[contains(@class, "taskButton")]'
RUBRIC_TABLE_XPATH      = '//table[contains(@class, "rubricView")]'
STUDENT_NAME_XPATH      = '//div[@class="checkbox"]/label/div/a'

# Spreadsheet Constants
HEADER_ROW = 6
RUBRIC_ROW = HEADER_ROW + 1

def populate_spreadsheet(assignment_name, assignments={}):

    print('Populating spreadsheet for %s...' % assignment_name)
    datapath = 'assignments/%s/Data' % (assignment_name)

    with open('%s/rubric.json' % datapath, 'r') as file:
        rubric = json.load(file)

    with open('%s/weights.json' % datapath, 'r') as file:
        weights = json.load(file)

    workbook_path = 'assignments/%s/grades.xlsx' % (assignment_name)
    wb = load_workbook(GRADING_TEMPLATE_PATH)

    wb.template = False
    ws = wb.active

    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    start_row = RUBRIC_ROW + 1
    end_row = len(assignments) + start_row - 1

    # Get spreadsheet header row values
    head   = list(ws.iter_rows())[HEADER_ROW - 1]
    header = {cell.internal_value : idx for idx, cell in enumerate(head, 1) if cell.internal_value}
    maxcol = max(header.values())

    # Format cells
    for i in range(start_row, end_row + 1):
        for c in ALPHABET[:maxcol]:
            ws['%s%s' % (c, i)].font = Font(name='Arial', size=12)
            ws['%s%s' % (c, i)].alignment = Alignment(horizontal='center', vertical='center')
        ws['%s%s' % (ALPHABET[header['Student']-1], i)].alignment = Alignment(horizontal='left')

    # Insert statistics
    totals_fmt = '%(col)s%(row_s)s:%(col)s%(row_e)s' % {'col'  :ALPHABET[header['Total']-1], 
                                                        'row_s':start_row,
                                                        'row_e':end_row}
    ws['A1'] = 'Generated on %s'   % (datetime.datetime.now())
    ws['B4'] = '=AVERAGE(%s)' % totals_fmt
    ws['C4'] = '=MEDIAN(%s)'  % totals_fmt
    ws['D4'] = '=STDEV(%s)'   % totals_fmt

    # Insert rubric questions
    for i, r in enumerate(rubric, 1):
        ws.cell(row=RUBRIC_ROW, column=header['Question %i'%i], value=r)

    # Insert task values
    sum_partial  = 'SUM(%(qone)s%(cr)s:%(qeight)s%(cr)s)'
    sum_template = '=+IF(%(sp)s=0,"",%(sp)s)' % {'sp':sum_partial} 
    sum_values   = {'qone'  :ALPHABET[header['Question 1']-1],
                    'qeight':ALPHABET[header['Question 8']-1]}
    for current_row, assignment in enumerate(assignments, start_row):
        sum_values.update({'cr' : current_row})

        ws.cell(row=current_row, column=header['Student'], value=assignment['name'].title())
        ws.cell(row=current_row, column=header['Total'], value=sum_template % sum_values)

        if assignment['name'] in weights:
            ws.cell(row=current_row, column=header['Weighted Score'], value=weights[assignment['name']])

        ws.cell(row=current_row, column=header['Word Count'], value=assignment['word_count'])
        ws.cell(row=current_row, column=header['Peer Feedback Link'], value=assignment['feedback_url'])
        ws.cell(row=current_row, column=maxcol+1, value=' ') # Make sure final value is not extended past col

    print('Saving grades to %s' % workbook_path)
    if os.path.exists(workbook_path):
        question = input("Do you wish to overwrite %s?  (y/n)" % (workbook_path))

        if question == 'y':
            wb.save(workbook_path)
        else:
            wb.save("assignments/%s/grades-new.xlsx" % assignment_name)
    else:
        wb.save(workbook_path)


def pull_assignments(session):
    """visits each assigned task, pulls the assignment as feedback_id"""
    resp = session.get(BASE_URL)
    page = resp.text
    tree = fromstring(page)

    assignment_name = tree.xpath(ASSIGNMENT_NAME_XPATH)[0].text.title()

    print('Pulling assignments for %s...' % (assignment_name))

    links = tree.xpath(ASSIGNMENT_LINKS_XPATH)

    if not os.path.exists('assignments/%s/Data' % (assignment_name)):
        os.makedirs('assignments/%s/Data' % (assignment_name))

    tasks = []
    rubric = []

    for curr_task_idx, link in enumerate(links, 1):
        st_name = None
        pf_link = link.get('href')
        fb_resp = session.get(BASE_URL + pf_link)
        fb_page = fb_resp.text
        fb_tree = fromstring(fb_page)
        dl_link = fb_tree.xpath(ASSIGNMENT_LINK_XPATH)[0].get('href')
        assert(dl_link)

        if not rubric:
            rubric_table = fb_tree.xpath(RUBRIC_TABLE_XPATH)[-1]
            rubric = [
                r.text.strip()
                for r in fb_tree.xpath(RUBRIC_TEXT_XPATH)]

        for a in fb_tree.xpath(STUDENT_NAME_XPATH):
            if a.text.strip():
                st_name = a.text.strip()
                break

        if st_name is None: assert(0), 'Couldn\'t pull student name'
        task = {
            'name': st_name.lower().strip(),
            'feedback_url': BASE_URL+pf_link,
            'feedback_id': pf_link.split('/')[-1],
            'paper_url': dl_link
        }

        filepath = 'assignments/%s/Papers/' % assignment_name
        if not os.path.exists(filepath):
            os.makedirs(filepath)
        filename = filepath + st_name.replace(' ', '') + '.pdf'
        if not os.path.exists(filename):
            resp = session.get(dl_link)
            with open(filename, 'wb') as f:
                f.write(resp.content)

        # Add word count to task info
        task['word_count'] = pdf_word_count(filename)
        tasks.append(task)
        print('%i/%i - %s' % (curr_task_idx, len(links), st_name))

    # Sort by student name
    tasks.sort(key=lambda k:k['name'])

    with open('assignments/%s/Data/assignments.json' % assignment_name, 'w') as file:
        json.dump(tasks, file)

    with open('assignments/%s/Data/rubric.json' % assignment_name, 'w') as file:
        json.dump(rubric, file)

    with open('assignments/%s/Data/weights.json' % assignment_name, 'w') as file:
        # Weighted scores reflect a 'best guess' score for the paper, based
        # on peerfeed back for the student. Still very rough at the moment,
        # and does poorly on outliers
        weights = get_weighted_scores(assignment_name, session)
        json.dump(weights, file)

    populate_spreadsheet(assignment_name, tasks)


def process():
    session = login()
    pull_assignments(session)

if __name__ == "__main__":
    process()
