# from google_spreadsheet import Sheet
from glob import glob
from os.path import exists
from openpyxl import load_workbook
from lxml.html import fromstring
from utils import login
import yaml, argparse

# The sheet id is found at the end of the spread sheet url
# If this is the url, for example:
#
#   https://docs.google.com/spreadsheets/d/1Ed_tyOHhyc-BAPKkcXJRD_R5xZbbb93aPYxNQl8wYbg/edit#gid=0
#
# the id is 1Ed_tyOHhyc-BAPKkcXJRD_R5xZbbb93aPYxNQl8wYbg
SHEET_ID = ''
BASE_URL = 'https://peerfeedback.gatech.edu'

RUBRIC_TABLE_XPATH  = '//table[contains(@class, "rubricView") and '+\
                         'not(contains(@id, "viewonly"))]'
RUBRIC_IDS_XPATH    = './/td[contains(@class, "rubric-element")]'
RUBRIC_ELE_ID_XPATH = 'data-rubric-element-combined-id'

def get_ta_name():
    try:
        with open('secrets.yml', 'r') as file:
            secrets = yaml.load(file)
            return secrets['spreadsheet-information']['ta-name']
    except:
        return ''


def get_grade_sheet(filename):
    wb = load_workbook(filename)
    ws = wb.worksheets[0]
    
    rows = list(ws.iter_rows())
    head = [cell.internal_value for cell in rows[5]]
    return [ {h:r.internal_value for h,r in zip(head, row)} for row in rows[7:] ]


def gs_submit(ta_name, rows, sheet_id):
    ''' Submit all grades in the grades.xlsx file to
        the google spreadsheet '''
    idxs = [0,2,3,4,5,6,7,8,9,11]
    data = [[r[i].internal_value if r[i].internal_value is not None else ''
            for i in idxs] for r in rows if r[0].internal_value]
    data = [[d[0]] + [ta_name] + [d[-1]] + d[1:-1] + [sum([int(v) if v else 0 for v in d[1:-1]])]
            for d in data]
    data = [[str(d) for d in row] for row in data]

    try:
        sheet = Sheet(sheet_id)
        sheet.write(data)
    except Exception as e:
        print(e) # Almost certainly occurs because of missing authentication file


def pf_submit(rows):
    session = login()

    for row in rows:
        if row['Comments']:
            scores = [int(row['Question %i'%i]) for i in range(1,9) if row['Question %i'%i]]
            data   = {'comment': str(row['Comments'])}
            print(row['Student'])

            url  = row['Peer Feedback Link']
            resp = session.get(url)
            page = resp.text
            tree = fromstring(page)

            table = tree.xpath(RUBRIC_TABLE_XPATH)
            assert(len(table) == 1), 'Multiple submission tables found..'

            tbl_rows = table[0].xpath('.//tr')
            assert(len(tbl_rows) == len(scores)), 'Different number of rubric items: '+\
                        'found %i on page; found %i in grading sheet' % (len(tbl_rows), len(scores))

            for i, r in enumerate(tbl_rows):
                ids = [td.get(RUBRIC_ELE_ID_XPATH) for td in
                        r.xpath(RUBRIC_IDS_XPATH)]
                assert(len(ids) == 5), 'Max %i criteria score found; should be 5?' % len(ids)

                data['rubricElements[%s]' % ids[scores[i] - 1]] = 'true'

            pfb_id = url.split('/')[-1]
            session.post(BASE_URL+'/drafts/%s/'%pfb_id, verify=False, data=data)
            

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--assignment', help="Which assignment to submit")
    parser.add_argument('--name', help="TA which program should submit for")
    parser.add_argument('--sheetid', help="Google spreadsheet id program should submit to")
    parser.add_argument('--submit_gs', action='store_true', help="Submit to the google spreadsheet instead")
    # parser.add_argument('--submit_pf', action='store_true', help="Whether to submit to peer feedback")
    args = parser.parse_args()

    submit_gs = bool(args.submit_gs)
    submit_pf = not submit_gs #bool(args.submit_pf)

    if not submit_gs and not submit_pf:
        raise Exception('Must set at least one of --submit_gs or --submit_pf flags')

    ta_name = get_ta_name()
    if not ta_name:
        ta_name = args.name if args.name else input('TA Name: ')

    assignment_input = args.assignment if args.assignment else input('Assigment to submit: ')

    if submit_gs:
        sheet_id = SHEET_ID
        if not SHEET_ID:
            sheet_id = args.sheetid if args.sheetid else input('Google Spreadsheet ID: ')


    assignment = None
    assignments= []
    for folder in glob('assignments/*/'):
        folder = folder.replace('\\','/')

        folder_assignment = folder.split('/')[1]
        assignments.append(folder_assignment.split('(')[0])
        if assignment_input.lower() in folder.lower():
            assignment = folder_assignment
            break
    if assignment is None:
        raise Exception('"%s" was not found; is the name correct?\n'%assignment_input +
                        'Available assignments are:\n\t- %s'%'\n\t- '.join(assignments))

    filename = './assignments/%s/grades.xlsx' % assignment
    if not exists(filename):
        raise Exception('%s does not exist - have you ran pull-assignments.py?'%filename)

    if submit_gs:
        resp = input('\nSubmit current grades to PF and master spreadsheet with vars:'
                    +'\n- Google spreadsheet: '+
                    '%s\n- Assignment: %s\n- TA: %s \ny/n?'%(sheet_id, assignment, ta_name))
        if resp != 'y': return

    rows = get_grade_sheet(filename)
    if submit_gs:    gs_submit(ta_name, rows, sheet_id)
    if submit_pf:    pf_submit(rows)





if __name__ == '__main__': main()

